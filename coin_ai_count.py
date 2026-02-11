import os
import cv2
import numpy as np
import tensorflow as tf
import json
from tensorflow.keras.preprocessing.image import ImageDataGenerator
from tensorflow.keras.applications import MobileNetV2
from tensorflow.keras.callbacks import ModelCheckpoint, EarlyStopping
from openpyxl import Workbook, load_workbook

# =======================
#   SETTINGS
# =======================
DATA_DIR = "coins_clean"
CLASSES = ["1tl", "5tl", "50kurus", "25kurus"]
IMG_SIZE = 224
BATCH_SIZE = 8
EPOCHS = 40
MODEL_OUT = "coins_model.h5"
CLASS_INDICES_FILE = "class_indices.json"
MAX_PER_CLASS = 60
CAMERA_URL = "http://192.168"
# =======================


def init_excel(file="results.xlsx"):
    if not os.path.exists(file):
        wb = Workbook()
        ws = wb.active
        ws.title = "Detections"
        ws.append(["ID", "Date - Time", "Total Money (TL)"])
        wb.save(file)
    return file


def append_excel(total, file="results.xlsx"):
    wb = load_workbook(file)
    ws = wb.active

    row_id = ws.max_row
    from datetime import datetime
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    ws.append([row_id, timestamp, total])
    wb.save(file)


for c in CLASSES:
    os.makedirs(os.path.join(DATA_DIR, c), exist_ok=True)


# ===============================================================
#  BACKGROUND REMOVAL (WHITE BACKGROUND -> COIN MASK)
# ===============================================================
def remove_background_from_frame(frame):
    gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
    gray = cv2.GaussianBlur(gray, (5, 5), 1)

    edges = cv2.Canny(gray, 40, 120)
    kernel = np.ones((5, 5), np.uint8)
    edges = cv2.dilate(edges, kernel, iterations=2)

    contours, _ = cv2.findContours(edges, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
    if not contours:
        return None

    cnt = max(contours, key=cv2.contourArea)
    if cv2.contourArea(cnt) < 3000:
        return None

    x, y, w, h = cv2.boundingRect(cnt)
    coin = frame[y:y+h, x:x+w]

    mask = np.zeros((h, w), dtype=np.uint8)
    cv2.drawContours(mask, [cnt - [x, y]], -1, 255, -1)

    coin_nobg = cv2.bitwise_and(coin, coin, mask=mask)

    canvas = np.ones_like(coin_nobg, dtype=np.uint8) * 255
    m = mask.astype(bool)
    canvas[m] = coin_nobg[m]

    return canvas


# ===============================================================
#  DATA COLLECTION
# ===============================================================
def collect_photos():
    print("\n=== Data Collection Mode ===")
    print("Classes:", CLASSES)
    print("Photos per class:", MAX_PER_CLASS)
    print("ENTER → Save, Q → Quit")

    cap = cv2.VideoCapture(CAMERA_URL)

    for label in CLASSES:
        print(f"\n*** Capture photos for {label} ***")
        save_dir = os.path.join(DATA_DIR, label)
        existing = len(os.listdir(save_dir))
        count = existing

        while count < MAX_PER_CLASS:
            ret, frame = cap.read()
            if not ret:
                break

            coin = remove_background_from_frame(frame)
            preview = frame.copy()

            if coin is not None:
                h, w = coin.shape[:2]
                preview[0:h, 0:w] = coin
                cv2.putText(preview, "COIN FOUND ✔", (10, h + 30),
                            cv2.FONT_HERSHEY_SIMPLEX, 1, (0,255,0), 2)
            else:
                cv2.putText(preview, "NO COIN", (10, 40),
                            cv2.FONT_HERSHEY_SIMPLEX, 1, (0,0,255), 2)

            cv2.putText(preview, f"{label}: {count}/{MAX_PER_CLASS}",
                        (10, preview.shape[0]-20),
                        cv2.FONT_HERSHEY_SIMPLEX, 1, (255,255,0), 2)

            cv2.imshow("Dataset Collection", preview)

            key = cv2.waitKey(1) & 0xFF
            if key == 13:
                if coin is not None:
                    filename = os.path.join(save_dir, f"{count}.jpg")
                    cv2.imwrite(filename, coin)
                    print("Saved:", filename)
                    count += 1
                else:
                    print("No coin detected!")
            if key == ord("q"):
                cap.release()
                cv2.destroyAllWindows()
                return

    cap.release()
    cv2.destroyAllWindows()
    print("\n*** Data collection completed! ***")


# ===============================================================
#  TRAIN MODEL
# ===============================================================
def train_model():
    print("\n=== Training Started ===")

    train_datagen = ImageDataGenerator(
        rescale=1/255.0,
        rotation_range=30,
        width_shift_range=0.2,
        height_shift_range=0.2,
        zoom_range=0.2,
        brightness_range=[0.6,1.4],
        validation_split=0.2
    )

    train_gen = train_datagen.flow_from_directory(
        DATA_DIR,
        target_size=(IMG_SIZE, IMG_SIZE),
        batch_size=BATCH_SIZE,
        class_mode="categorical",
        subset="training"
    )

    val_gen = train_datagen.flow_from_directory(
        DATA_DIR,
        target_size=(IMG_SIZE, IMG_SIZE),
        batch_size=BATCH_SIZE,
        class_mode="categorical",
        subset="validation"
    )

    class_idx = train_gen.class_indices
    idx_to_label = {str(v): k for k, v in class_idx.items()}
    json.dump(idx_to_label, open(CLASS_INDICES_FILE,"w"), indent=2)

    base = MobileNetV2(include_top=False, input_shape=(IMG_SIZE,IMG_SIZE,3), weights="imagenet")
    base.trainable = True

    model = tf.keras.Sequential([
        base,
        tf.keras.layers.GlobalAveragePooling2D(),
        tf.keras.layers.Dense(256, activation="relu"),
        tf.keras.layers.Dropout(0.4),
        tf.keras.layers.Dense(len(class_idx), activation="softmax")
    ])

    model.compile(optimizer=tf.keras.optimizers.Adam(1e-4),
                  loss="categorical_crossentropy",
                  metrics=["accuracy"])

    checkpoint = ModelCheckpoint(MODEL_OUT, save_best_only=True, monitor="val_accuracy")
    early = EarlyStopping(patience=5, restore_best_weights=True)

    model.fit(train_gen, validation_data=val_gen,
              epochs=EPOCHS,
              callbacks=[checkpoint, early])

    model.save(MODEL_OUT)
    print("\n=== Model Trained ===")


# ===============================================================
#  REAL-TIME DETECTION
# ===============================================================
def realtime_detect():
    excel_file = init_excel()
    last_saved_total = None

    if not os.path.exists(MODEL_OUT):
        print("Model not found! Train first.")
        return

    model = tf.keras.models.load_model(MODEL_OUT)

    with open(CLASS_INDICES_FILE, "r") as f:
        idx_to_label = json.load(f)
    idx_to_label = {int(k): v for k, v in idx_to_label.items()}

    values = {
        "1tl": 1.0,
        "5tl": 5.0,
        "50kurus": 0.50,
        "25kurus": 0.25
    }

    cap = cv2.VideoCapture(CAMERA_URL)
    cap.set(cv2.CAP_PROP_BUFFERSIZE, 1)

    print("Real-time multi-coin detection started (q = quit).")

    SKIP = 4
    frame_count = 0
    last_results = []
    last_total = 0.0

    while True:
        ret, frame = cap.read()
        if not ret:
            continue

        display = frame.copy()
        frame_count += 1

        if frame_count % SKIP == 0:
            gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
            gray = cv2.GaussianBlur(gray, (5, 5), 1)

            _, thresh = cv2.threshold(gray, 0, 255,
                                      cv2.THRESH_BINARY_INV + cv2.THRESH_OTSU)

            contours, _ = cv2.findContours(thresh, cv2.RETR_EXTERNAL,
                                           cv2.CHAIN_APPROX_SIMPLE)

            results = []
            total_money = 0.0

            for cnt in contours:
                if cv2.contourArea(cnt) < 1500:
                    continue

                x, y, w, h = cv2.boundingRect(cnt)
                coin = frame[y:y+h, x:x+w]

                mask = np.zeros((h, w), dtype=np.uint8)
                cv2.drawContours(mask, [cnt - [x, y]], -1, 255, -1)

                coin_nobg = cv2.bitwise_and(coin, coin, mask=mask)

                resized = cv2.resize(coin_nobg, (IMG_SIZE, IMG_SIZE)) / 255.0
                resized = np.expand_dims(resized, 0)

                pred = model.predict(resized, verbose=0)[0]
                idx = int(np.argmax(pred))
                cls = idx_to_label[idx]
                prob = float(pred[idx])
                val = values.get(cls, 0)

                results.append((x, y, w, h, cls, prob))
                total_money += val

            last_results = results
            last_total = total_money

        for (x, y, w, h, cls, prob) in last_results:
            cv2.rectangle(display, (x, y), (x+w, y+h), (0,255,0), 2)
            cv2.putText(display, f"{cls} {prob:.2f}",
                        (x, y-10), cv2.FONT_HERSHEY_SIMPLEX,
                        0.7, (0,255,0), 2)

        cv2.putText(display, f"TOTAL: {last_total:.2f} TL",
                    (10, 40),
                    cv2.FONT_HERSHEY_SIMPLEX,
                    1.2, (0,255,255), 3)

        if last_saved_total != last_total:
            append_excel(last_total)
            last_saved_total = last_total
            print(f"[EXCEL] Saved: {last_total} TL")

        cv2.imshow("Multi-Coin Detect", display)

        if cv2.waitKey(1) & 0xFF == ord("q"):
            break

    cap.release()
    cv2.destroyAllWindows()


# ===============================================================
#  MENU
# ===============================================================
while True:
    print("\n===============================")
    print("   COINN")
    print("===============================")
    print("1 - Collect Data")
    print("2 - Train Model")
    print("3 - Real-Time Detection")
    print("0 - Exit")

    sec = input("Select: ")

    if sec == "1":
        collect_photos()
    elif sec == "2":
        train_model()
    elif sec == "3":
        realtime_detect()
    elif sec == "0":
        break
    else:
        print("Invalid selection!")
