import cv2
import numpy as np
from scipy import ndimage
from openpyxl import Workbook, load_workbook
from datetime import datetime
import os

EXCEL_FILE = "money_records.xlsx"

# Create Excel file if it does not exist
if not os.path.exists(EXCEL_FILE):
    wb = Workbook()
    ws = wb.active
    ws.title = "Records"
    ws.append([
        "Date",
        "Time",
        "Total Money (TL)"
    ])
    wb.save(EXCEL_FILE)

wb = load_workbook(EXCEL_FILE)
ws = wb.active

last_total = None

CAMERA_URL = "http://172.18"
cap = cv2.VideoCapture(CAMERA_URL)

while True:
    ret, image_original = cap.read()
    if not ret:
        break

    image = cv2.cvtColor(image_original, cv2.COLOR_BGR2GRAY)

    # Automatic thresholding using OTSU method
    # Purpose: separate coins from background
    _, bw = cv2.threshold(
        image,
        0,
        255,
        cv2.THRESH_BINARY + cv2.THRESH_OTSU
    )

    # Invert (coins white, background black)
    bw = cv2.bitwise_not(bw)

    # Fill holes inside coins (due to glare or shadows)
    bw = ndimage.binary_fill_holes(bw).astype(np.uint8) * 255

    # Find connected components (objects)
    num_labels, labels, stats, _ = cv2.connectedComponentsWithStats(bw)

    # Create clean mask
    cleaned = np.zeros_like(bw)

    # Check each object
    for i in range(1, num_labels):
        # Remove very small objects (dust, noise, reflections)
        if stats[i, cv2.CC_STAT_AREA] >= 30:
            cleaned[labels == i] = 255

    # Create elliptical structuring element (kernel)
    kernel = cv2.getStructuringElement(
        cv2.MORPH_ELLIPSE, (23, 23)
    )

    # Erosion:
    # - Smooths coin edges
    # - Removes thin connections
    bw2 = cv2.erode(cleaned, kernel)

    # ========================================================
    # CONTOUR DETECTION (ONE SHAPE PER COIN)
    # ========================================================
    contours, _ = cv2.findContours(
        bw2,
        cv2.RETR_EXTERNAL,       # Only external contours
        cv2.CHAIN_APPROX_SIMPLE
    )

    total = 0  # Total money for this frame

    for cnt in contours:
        area = cv2.contourArea(cnt)

        # Ignore very small contours
        if area < 200:
            continue

        # Moments (for center calculation)
        M = cv2.moments(cnt)
        if M["m00"] == 0:
            continue

        # Center of the coin
        cx = int(M["m10"] / M["m00"])
        cy = int(M["m01"] / M["m00"])

        # Scale area (empirical adjustment)
        area = area / 100

        if area > 720:
            total += 5
            label = "5 TL"

        elif 600 < area < 710:
            total += 1
            label = "1 TL"

        elif 500 < area < 600:
            total += 0.5
            label = "50 Kr"

        elif 350 < area < 450:
            total += 0.25
            label = "25 Kr"

        elif 300 < area < 350:
            total += 0.10
            label = "10 Kr"

        else:
            continue

        # Write coin value on image
        cv2.putText(
            image_original,
            label,
            (cx - 30, cy - 10),
            cv2.FONT_HERSHEY_SIMPLEX,
            0.6,
            (0, 255, 0),
            2
        )

        cv2.putText(
            image_original,
            f"Area: {int(area)}",
            (cx - 30, cy + 20),
            cv2.FONT_HERSHEY_SIMPLEX,
            0.5,
            (255, 255, 255),
            2
        )

        # Draw coin contour
        cv2.drawContours(
            image_original, [cnt],
            -1, (0, 255, 255), 2
        )

    # Save only if total value has changed
    if last_total != total:
        now = datetime.now()
        date = now.strftime("%Y-%m-%d")
        time = now.strftime("%H:%M:%S")

        ws.append([
            date,
            time,
            round(total, 2)
        ])

        wb.save(EXCEL_FILE)
        last_total = total

    cv2.putText(
        image_original,
        f"Total Money: {total:.2f} TL",
        (20, 40),
        cv2.FONT_HERSHEY_SIMPLEX,
        1, (0, 0, 255), 3
    )

    cv2.imshow("Coin Detection - Live", image_original)
    cv2.imshow("Mask", bw2)

    # Press 'q' to quit
    if cv2.waitKey(1) & 0xFF == ord('q'):
        break

cap.release()
wb.save(EXCEL_FILE)
cv2.destroyAllWindows()
